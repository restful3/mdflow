"""xlsx -> Markdown via openpyxl.

Loaded read_only + data_only (memory-safe; formula cells yield their last
cached value). Each sheet renders as `## <SheetName>` plus a Markdown
table over the used range, first row as header. Empty sheets render the
heading plus `(empty sheet)`. No internal try/except.
"""

from __future__ import annotations

import io

from openpyxl import load_workbook

from mdflow.converters._md_table import escape_table_cell
from mdflow.converters.base import (
    ConversionContext,
    ConversionResult,
    ProgressCallback,
)


class XlsxConverter:
    name = "xlsx-openpyxl"
    formats: tuple[str, ...] = ("xlsx",)
    requires_gpu = False

    def can_handle(self, ctx: ConversionContext) -> bool:
        return ctx.format in self.formats

    def convert(self, ctx: ConversionContext, progress: ProgressCallback) -> ConversionResult:
        progress("parse", 10)
        wb = load_workbook(io.BytesIO(ctx.data), data_only=True, read_only=True)
        try:
            names = wb.sheetnames
            total = max(len(names), 1)
            blocks: list[str] = []
            for i, name in enumerate(names, start=1):
                blocks.append(_sheet_to_md(name, wb[name]))
                progress("render", 10 + int(80 * i / total))
        finally:
            wb.close()
        markdown = "\n\n".join(blocks).strip()
        progress("done", 100)
        return ConversionResult(markdown=markdown, metadata={"formula_values": "cached"})


def _cell(value) -> str:
    return "" if value is None else escape_table_cell(str(value))


def _sheet_to_md(name: str, ws) -> str:
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    while rows and all(c is None for c in rows[-1]):
        rows.pop()
    if not rows:
        return f"## {name}\n\n(empty sheet)"

    width = max(len(r) for r in rows)
    norm = [[_cell(v) for v in r] + [""] * (width - len(r)) for r in rows]
    header, *body = norm
    table = ["| " + " | ".join(header) + " |"]
    table.append("| " + " | ".join("---" for _ in header) + " |")
    for r in body:
        table.append("| " + " | ".join(r) + " |")
    return f"## {name}\n\n" + "\n".join(table)
